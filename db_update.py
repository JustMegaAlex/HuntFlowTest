import sys
import os
import openpyxl
import argparse
import re
from api import API

# handle external arguments
argparser = argparse.ArgumentParser()
argparser.add_argument('--token', '-t', required = True)
argparser.add_argument('--path', '-p', required = True)

ARGS = argparser.parse_args()  

STATUSES_MAPPING = {
    'Отправлено письмо': 'Contacted',
    'Интервью с HR': 'HR Interview',
    'Выставлен оффер': 'Offered',
    'Отказ': 'Declined'
}

USELOGS = False

SUCCESS_MESS = 'Succefully added all applicants!'

PROCEEDED = []

DBFILE = 'Тестовая база.xlsx'

api = API(ARGS.token)

STATUSES_IDS_MAPPING = api.get_statuses_ids_mapping()

VACANCIES_IDS_MAPPING = api.get_vacancies_ids_mapping()

QUOTAS_IDS_MAPPING = api.get_vacancies_quotas_ids_mapping(VACANCIES_IDS_MAPPING.values())


def load_candidates_data(path):

    field_names = {'position':1, 'name':2, 'money':3, 'comment':4, 'status_name':5}
    data = []
    path = fr'{path}'
    xl_path = os.path.join(path, DBFILE)
    wb = openpyxl.load_workbook(xl_path)
    ws = wb.active
    row = 2
    cell_val = ws.cell(row, 1).value

    while cell_val:

        cand_data = {}

        status_name = ws.cell(row, field_names['status_name']).value
        status_api_name = STATUSES_MAPPING[status_name]
        name = ws.cell(row, field_names['name']).value
        name_list = name.split()
        position = ws.cell(row, field_names['position']).value
        file_path = get_resume_local_path(path, name, position)

        cand_data['position'] = position
        cand_data['vacancy'] = VACANCIES_IDS_MAPPING[position]
        cand_data['money'] = ws.cell(row, field_names['money']).value
        cand_data['comment'] = ws.cell(row, field_names['comment']).value
        cand_data['status'] = STATUSES_IDS_MAPPING[status_api_name]
        cand_data['second_name'] = name_list[0]
        cand_data['first_name'] = name_list[1]
        cand_data['middle_name'] = name_list[2] if len(name_list) == 3 else ''
        # add resume file path if exists
        cand_data['local_file'] = file_path

        cand_data['rejection_reason'] = cand_data['comment'] if status_api_name == 'Declined' else None

        data.append(cand_data)
        row += 1
        cell_val = ws.cell(row, 1).value

    return data

def get_resume_local_path(db_path, name, position):

    files_path = os.path.join(db_path, position)
    
    files = os.listdir(files_path)

    resume_path = None

    for f in files:

        # fix wrong 'й' encoding
        f_check = re.sub(b'\xd0\xb8\xcc\x86'.decode('utf8'), 'й', f)

        if name.strip() in f_check.strip():

            resume_path = os.path.join(files_path, f)

            break

    return resume_path

def create_cand_db_data(src_data):
    
    cand_data = {
        'last_name': src_data['second_name'],
        'first_name': src_data['first_name'],
        'middle_name': src_data['middle_name'],
        'phone': src_data['phone'],
        'email': src_data['email'],
        'position': src_data['position'],
        'company': None,
        'money': src_data['money'],
        'birthday_day': src_data['birthday_day'],
        'birthday_month': src_data['birthday_month'],
        'birthday_year': src_data['birthday_year'],
        'photo': src_data['photo'],
        'externals': src_data['externals'],
        # 'rejection_reason': src_data['rejection_reason'], # don't know where this should be taken from, couses 400 status code if send as string 
        'vacancy': src_data['vacancy'],
        'status': src_data['status'],
        'comment': src_data['comment'],
    }

    return cand_data

def create_cand_vacancy_data(src_data):

    cand_data = {
        'id': src_data['id'],
        'vacancy': src_data['vacancy'],
        'status': src_data['status'],
        'comment': src_data['comment'],
        'files': [
            {
                'id': src_data['externals'][0]['files'][0]['id'],
            }
        ],
        # 'rejection_reason': src_data['rejection_reason'],
    }

    if cand_data['status'] == STATUSES_IDS_MAPPING['Hired']:

        cand_data['fill_quota'] = QUOTAS_IDS_MAPPING[src_data['vacancy']]

    return cand_data

def log(mess):

    with open('logs.txt', 'a', encoding = 'utf8') as f:

            f.write(mess + '\n')

    print(mess)

def clear_log():

    with open('logs.txt', 'w'):

        pass


if __name__ == '__main__':

    # check logs
    if os.path.exists('logs.txt'):

        with open('logs.txt', encoding = 'utf8') as f:

            logs_text = f.read()

        # if last launch was successful clear logs 
        if re.search(rf'{SUCCESS_MESS}', logs_text):

            clear_log()
        
        # otherwise decide whether to use logs or not
        elif logs_text:

            print('log file found, should I use it?')

            ans = ''

            while ans not in ['y', 'n']:

                ans = input('please, type \'y\' or \'n\': ')

            if ans == 'y':

                proceeded_names = re.findall(r'(?<=\<).+?(?=\>)', logs_text)

                PROCEEDED = [tuple(name.split()) for name in proceeded_names]

            else:

                clear_log()

    data = load_candidates_data(ARGS.path)

    for cand in data:

        first_name, last_name = cand['first_name'], cand['second_name']

        # check if applpicant is already added
        if (last_name, first_name) in PROCEEDED:

            print(f'Applicant <{first_name} {last_name}> has already been added to HuntFlow')

            continue

        data_from_file = api.upload_resume(cand['local_file'])

        fields = data_from_file.get('fields')
        birth_date = fields.get('birthdate')
        phones = fields.get('phones')
        phone_num = phones[0] if phones else None

        externals = [{
            'data': {
                'body': data_from_file.get('text')
            },
            'auth_type': 'NATIVE',
            'files': [
                {
                    'id': data_from_file.get('id')
                }
            ],
            'account_source': None
        }]

        cand['resume_text'] = data_from_file.get('text')
        cand['email'] = fields.get('email')
        cand['birthday_day'] = birth_date.get('day') if birth_date else None
        cand['birthday_month'] = birth_date.get('month') if birth_date else None
        cand['birthday_year'] = birth_date.get('year') if birth_date else None
        cand['phone'] = phone_num
        cand['photo'] = data_from_file.get('photo').get('id')
        cand['externals'] = externals

        cand = create_cand_db_data(cand)

        db_data = api.add_candidate(cand)

        cand.update({'id':db_data['id']})

        vac_data = create_cand_vacancy_data(cand)

        cand_vacancy_info = api.add_vacancy_candidate(vac_data)

        log(f'Applicant <{last_name} {first_name}> added to a vacancy')
        
    log(SUCCESS_MESS)